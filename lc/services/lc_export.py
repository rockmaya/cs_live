from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from datetime import datetime

def export_lcs_to_excel(computed_lcs):
    wb = Workbook()
    ws = wb.active
    ws.title = "LC List"

    headers = [
        "Bank Name",
        "Swift Code",
        "Global Limit",
        "LC No",
        "Opening Date",
        "LC Amount",
        "CCY",
        "Remaining Amount",
        "Maturity Date",
        "Matured In",
        "Status",
        "Created By",
        "Created At",
        "Updated By",
        "Updated At",
    ]

    ws.append(headers)

    for row in computed_lcs:
        lc = row["lc"]
        ws.append([
            lc.bank.name,
            lc.swift_code,
            float(lc.global_limit),
            lc.lc_no,
            lc.opening_date,
            float(lc.lc_amount),
            "USD",
            float(row["remaining_amount"]),
            lc.maturity_date,
            row["matured_in"],
            lc.created_by.username if lc.created_by else "",
            lc.created_at.strftime("%Y-%m-%d %H:%M"),
            lc.updated_by.username if lc.updated_by else "",
            lc.updated_at.strftime("%Y-%m-%d %H:%M"),
        ])

    # Auto-size columns
    for col in ws.columns:
        ws.column_dimensions[get_column_letter(col[0].column)].width = 18

    return wb
